import os
from django.core.management.base import BaseCommand
from django.db import transaction
from openpyxl import load_workbook
from usuario.models import Funcionario, Cargo

class Command(BaseCommand):
    help = 'Importa cargos de um arquivo Excel (XLSX) e vincula aos funcionários pelas matrículas'

    def add_arguments(self, parser):
        parser.add_argument(
            '--file',
            type=str,
            help='Caminho completo para o arquivo Excel (Cargos.xlsx)',
        )

    def handle(self, *args, **options):
        # Obtém o caminho do arquivo do argumento ou usa o padrão
        file_path = options.get('file') or 'Cargos.xlsx'
        
        if not os.path.exists(file_path):
            self.stdout.write(self.style.ERROR(f'Arquivo não encontrado: {file_path}'))
            return

        try:
            wb = load_workbook(filename=file_path)
            sheet = wb.active
            
            # Contadores para estatísticas
            total_linhas = 0
            cargos_criados = 0
            funcionarios_atualizados = 0
            funcionarios_nao_encontrados = []
            
            with transaction.atomic():
                for row in sheet.iter_rows(min_row=2, values_only=True):  # Pula cabeçalho
                    if not row[0]:  # Se a primeira célula estiver vazia, pula a linha
                        continue
                        
                    total_linhas += 1
                    
                    codigo = int(row[0]) if row[0] else None
                    nome_funcionario = str(row[1]).strip() if row[1] else None
                    nome_cargo = str(row[2]).strip() if row[2] else None
                    
                    if not all([codigo, nome_funcionario, nome_cargo]):
                        self.stdout.write(self.style.WARNING(
                            f'Linha {total_linhas+1} incompleta - ignorada'
                        ))
                        continue
                    
                    try:
                        # Tenta converter o código para inteiro (matrícula)
                        matricula = int(codigo)
                    except ValueError:
                        self.stdout.write(self.style.WARNING(
                            f'Matrícula inválida na linha {total_linhas+1}: {codigo}'
                        ))
                        continue
                    
                    # Busca ou cria o cargo (case insensitive)
                    nome_cargo = nome_cargo.upper()  # Padroniza para maiúsculas
                    cargo, created = Cargo.objects.get_or_create(
                        nome__iexact=nome_cargo,
                        defaults={'nome': nome_cargo}
                    )
                    
                    if created:
                        cargos_criados += 1
                        self.stdout.write(self.style.SUCCESS(
                            f'Cargo criado: {nome_cargo}'
                        ))
                    
                    # Atualiza o funcionário
                    try:
                        funcionario = Funcionario.objects.get(matricula=matricula)
                        if funcionario.cargo != cargo:
                            funcionario.cargo = cargo
                            funcionario.save()
                            funcionarios_atualizados += 1
                            self.stdout.write(self.style.SUCCESS(
                                f'Funcionário atualizado: {matricula} - {nome_funcionario}'
                            ))
                    except Funcionario.DoesNotExist:
                        funcionarios_nao_encontrados.append(
                            f"{matricula} - {nome_funcionario}"
                        )
                        self.stdout.write(self.style.WARNING(
                            f'Funcionário não encontrado: {matricula} - {nome_funcionario}'
                        ))
            
            # Exibe estatísticas finais
            self.stdout.write(self.style.SUCCESS('\n' + '='*50))
            self.stdout.write(self.style.SUCCESS(
                f'RELATÓRIO FINAL:\n'
                f'Total de linhas processadas: {total_linhas}\n'
                f'Cargos criados: {cargos_criados}\n'
                f'Funcionários atualizados: {funcionarios_atualizados}\n'
                f'Funcionários não encontrados: {len(funcionarios_nao_encontrados)}'
            ))
            
            if funcionarios_nao_encontrados:
                self.stdout.write(self.style.WARNING(
                    '\nLista de funcionários não encontrados:'
                ))
                for func in funcionarios_nao_encontrados:
                    self.stdout.write(self.style.WARNING(f'- {func}'))
            
        except Exception as e:
            self.stdout.write(self.style.ERROR(f'Erro durante o processamento: {str(e)}'))
            raise e