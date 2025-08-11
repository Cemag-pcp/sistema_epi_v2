from django.shortcuts import render
from django.http import JsonResponse, HttpResponse
from django.db.models import Sum
from solicitacao.models import Solicitacao
from devolucao.models import Devolucao
from usuario.models import Funcionario

from PIL import Image
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as ExcelImage

from io import BytesIO
from datetime import timedelta
import os
import base64
import copy

# Create your views here.
def template_ficha(request):

    funcionarios = Funcionario.objects.values('id','matricula','nome').order_by('id')

    return render(request, 'ficha.html', {'funcionarios':funcionarios})

def gerar_ficha_epi(request, id):
    try:
        # 1. Obter dados do funcionário
        funcionario = Funcionario.objects.get(id=id)
        
        # 2. Obter solicitações e devoluções
        solicitacoes = Solicitacao.objects.filter(
            funcionario=funcionario,
            status='Entregue'
        ).prefetch_related(
            'dados_solicitacao',
            'dados_solicitacao__dados_solicitacao_devolucao',
            'assinatura'
        ).order_by('data_solicitacao')

        # Criar lista combinada de itens
        itens_planilha = []
        
        for solicitacao in solicitacoes:
            for dado in solicitacao.dados_solicitacao.all():
                itens_planilha.append({
                    'tipo': 'solicitacao',
                    'objeto': solicitacao,
                    'dado': dado,
                    'data': solicitacao.data_solicitacao
                })
                
                for devolucao in dado.dados_solicitacao_devolucao.all():
                    itens_planilha.append({
                        'tipo': 'devolucao',
                        'objeto': devolucao,
                        'dado': dado,
                        'data': devolucao.data_devolucao
                    })

        # Ordenar itens por data
        itens_planilha.sort(key=lambda x: x['data'])
        
        # 3. Preparar planilha
        wb = load_workbook('FICHA DE EPI.xlsx')
        ws = wb.active
        
        # Preencher cabeçalho
        ws['B4'] = funcionario.nome
        ws['B5'] = funcionario.matricula
        ws['B6'] = funcionario.cargo.nome if funcionario.cargo else ""
        ws['B7'] = funcionario.data_admissao.strftime("%d/%m/%Y") if funcionario.data_admissao else ""
        ws['B8'] = funcionario.setor.nome if funcionario.setor else ""

        # 4. Processar itens
        temp_files = []
        assinatura_adicionada_b22 = False
        
        for i, item in enumerate(itens_planilha):
            linha_destino = 27 + i
            
            # Copiar formatação da linha modelo (27)
            for coluna in range(1, 9):
                fonte = copy.copy(ws.cell(row=27, column=coluna).font)
                preenchimento = copy.copy(ws.cell(row=27, column=coluna).fill)
                borda = copy.copy(ws.cell(row=27, column=coluna).border)
                alinhamento = copy.copy(ws.cell(row=27, column=coluna).alignment)
                
                celula = ws.cell(row=linha_destino, column=coluna)
                celula.font = fonte
                celula.fill = preenchimento
                celula.border = borda
                celula.alignment = alinhamento
            
            ws.row_dimensions[linha_destino].height = ws.row_dimensions[27].height
            
            if item['tipo'] == 'solicitacao':
                solicitacao = item['objeto']
                dado = item['dado']
                data_formatada = (solicitacao.data_solicitacao - timedelta(hours=3)).strftime("%d/%m/%Y %H:%M")
                
                ws.cell(row=linha_destino, column=1).value = data_formatada
                ws.cell(row=linha_destino, column=2).value = dado.quantidade
                ws.cell(row=linha_destino, column=3).value = f"{dado.equipamento.codigo} - {dado.equipamento.nome}"
                ws.cell(row=linha_destino, column=4).value = dado.equipamento.ca
                ws.cell(row=linha_destino, column=6).value = dado.get_motivo_display()
                
                if hasattr(solicitacao, 'assinatura') and solicitacao.assinatura.imagem_assinatura:
                    # Criar arquivo temporário
                    temp_filename = f"assinatura_temp_{solicitacao.id}_{i}.png"
                    with open(temp_filename, 'wb') as temp_file:
                        solicitacao.assinatura.imagem_assinatura.seek(0)
                        temp_file.write(solicitacao.assinatura.imagem_assinatura.read())
                    temp_files.append(temp_filename)
                    
                    # Adicionar imagem em G{linha_destino}
                    img_linha = ExcelImage(temp_filename)
                    img_linha.anchor = f'G{linha_destino}'
                    img_linha.height = 130
                    img_linha.width = 150
                    ws.add_image(img_linha)
                    
                    # Adicionar em B22 apenas na primeira vez (com NOVA instância)
                    if not assinatura_adicionada_b22:
                        img_b22 = ExcelImage(temp_filename)  # Nova instância
                        img_b22.anchor = 'B22'
                        img_b22.height = 130
                        img_b22.width = 150
                        ws.add_image(img_b22)
                        assinatura_adicionada_b22 = True

        # 5. Salvar e retornar o arquivo
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="Ficha_EPI_{funcionario.nome}.xlsx"'
        
        wb.save(response)
        wb.close()
        
        # Limpar arquivos temporários
        for temp_file in temp_files:
            try:
                if os.path.exists(temp_file):
                    os.remove(temp_file)
            except:
                pass
        
        return response

    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)

def processar_assinatura(base64_string, index):
    """Helper function to process signature images"""
    # Remove prefix if exists
    if isinstance(base64_string, str) and ';base64,' in base64_string:
        base64_string = base64_string.split(';base64,')[1]
    
    image_data = base64.b64decode(base64_string)
    image_buffer = BytesIO(image_data)
    image = Image.open(image_buffer)
    
    # Save temporary file
    temp_filename = f"assinatura{index}.png"
    image.save(temp_filename)
    
    # Create Excel image object
    img = ExcelImage(temp_filename)
    img.height = 130
    img.width = 150
    
    return img